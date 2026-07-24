#!/usr/bin/env python3
"""将日本球场原始 Excel 转换为三 Sheet 多语言标准工作簿。"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TOOL_DIR = Path(__file__).resolve().parent
DATA_DIR = TOOL_DIR / "data"
REQUIRED_COLUMNS = ["球场id中文", "国家", "球场名称中文", "城市", "地址"]
LANGUAGES = ["de", "fr", "es", "sv", "nl", "vi", "ms", "id"]

KO = {
    "ア":"아","イ":"이","ウ":"우","エ":"에","オ":"오","カ":"카","キ":"키","ク":"쿠","ケ":"케","コ":"코",
    "ガ":"가","ギ":"기","グ":"구","ゲ":"게","ゴ":"고","サ":"사","シ":"시","ス":"스","セ":"세","ソ":"소",
    "ザ":"자","ジ":"지","ズ":"즈","ゼ":"제","ゾ":"조","タ":"타","チ":"치","ツ":"쓰","テ":"테","ト":"토",
    "ダ":"다","ヂ":"지","ヅ":"즈","デ":"데","ド":"도","ナ":"나","ニ":"니","ヌ":"누","ネ":"네","ノ":"노",
    "ハ":"하","ヒ":"히","フ":"후","ヘ":"헤","ホ":"호","バ":"바","ビ":"비","ブ":"부","ベ":"베","ボ":"보",
    "パ":"파","ピ":"피","プ":"푸","ペ":"페","ポ":"포","マ":"마","ミ":"미","ム":"무","メ":"메","モ":"모",
    "ヤ":"야","ユ":"유","ヨ":"요","ラ":"라","リ":"리","ル":"루","レ":"레","ロ":"로","ワ":"와","ヲ":"오",
    "キャ":"캬","キュ":"큐","キョ":"쿄","シャ":"샤","シュ":"슈","ショ":"쇼","チャ":"차","チュ":"추","チョ":"초",
    "ニャ":"냐","ニュ":"뉴","ニョ":"뇨","ヒャ":"햐","ヒュ":"휴","ヒョ":"효","ミャ":"먀","ミュ":"뮤","ミョ":"묘",
    "リャ":"랴","リュ":"류","リョ":"료","ギャ":"갸","ギュ":"규","ギョ":"교","ジャ":"자","ジュ":"주","ジョ":"조",
    "ビャ":"뱌","ビュ":"뷰","ビョ":"뵤","ピャ":"퍄","ピュ":"퓨","ピョ":"표",
}
TH = {
    "ア":"อะ","イ":"อิ","ウ":"อุ","エ":"เอะ","オ":"โอ","カ":"คะ","キ":"คิ","ク":"คุ","ケ":"เคะ","コ":"โกะ",
    "ガ":"กะ","ギ":"กิ","グ":"กุ","ゲ":"เกะ","ゴ":"โกะ","サ":"ซะ","シ":"ชิ","ス":"ซุ","セ":"เซ","ソ":"โซะ",
    "ザ":"ซะ","ジ":"จิ","ズ":"ซุ","ゼ":"เซ","ゾ":"โซะ","タ":"ทะ","チ":"จิ","ツ":"สึ","テ":"เทะ","ト":"โตะ",
    "ダ":"ดะ","デ":"เดะ","ド":"โดะ","ナ":"นะ","ニ":"นิ","ヌ":"นุ","ネ":"เนะ","ノ":"โนะ","ハ":"ฮะ","ヒ":"ฮิ",
    "フ":"ฟุ","ヘ":"เฮะ","ホ":"โฮะ","バ":"บะ","ビ":"บิ","ブ":"บุ","ベ":"เบะ","ボ":"โบะ","パ":"พะ","ピ":"พิ",
    "プ":"พุ","ペ":"เพะ","ポ":"โปะ","マ":"มะ","ミ":"มิ","ム":"มุ","メ":"เมะ","モ":"โมะ","ヤ":"ยะ","ユ":"ยุ",
    "ヨ":"โยะ","ラ":"ระ","リ":"ริ","ル":"รุ","レ":"เระ","ロ":"โระ","ワ":"วะ","ヲ":"โอะ","キャ":"เคีย","キュ":"คิว",
    "キョ":"เคียว","シャ":"ชะ","シュ":"ชู","ショ":"โชะ","チャ":"จะ","チュ":"จู","チョ":"โจ","ニャ":"เนีย","ニュ":"นิว",
    "ニョ":"เนียว","ヒャ":"เฮีย","ヒュ":"ฮิว","ヒョ":"เฮียว","リャ":"เรีย","リュ":"ริว","リョ":"เรียว","ジャ":"จะ","ジュ":"จู","ジョ":"โจ",
}


def read_json(path: Path):
    if not path.exists():
        raise FileNotFoundError(f"缺少工具数据文件：{path.name}")
    return json.loads(path.read_text(encoding="utf-8"))


def load_reference():
    prefectures = read_json(DATA_DIR / "prefectures.json")
    municipalities = read_json(DATA_DIR / "municipalities.json")
    corrections = read_json(DATA_DIR / "source_corrections.json")
    pref_by_ja = {p["ja"]: p for p in prefectures}
    muni_by_key = {(m["prefecture"], m["name"]): m for m in municipalities}
    muni_by_pref = defaultdict(list)
    for item in municipalities:
        muni_by_pref[item["prefecture"]].append(item)
    return prefectures, pref_by_ja, muni_by_key, muni_by_pref, corrections


def compact(value) -> str:
    return re.sub(r"[\s\u3000\xa0]+", "", str(value or "")).replace("ヶ", "ケ").replace("ヵ", "カ")


def latin_compact(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").casefold())


def id_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def build_pref_aliases(prefectures):
    aliases = {}
    for pref in prefectures:
        for alias in {pref["ja"], pref["zh_cn"], pref["zh_tw"], pref["en"]}:
            aliases[compact(alias)] = pref["ja"]
            aliases[latin_compact(alias)] = pref["ja"]
    aliases.update({compact(k): v for k, v in {
        "琦玉县":"埼玉県", "枥木县":"栃木県", "鹿儿岛県":"鹿児島県", "兵库県":"兵庫県", "千叶県":"千葉県",
        "新泻县":"新潟県", "新泻県":"新潟県", "福冈县":"福岡県", "冲绳県":"沖縄県", "茨城":"茨城県", "分県":"大分県",
    }.items()})
    return aliases


def detect_prefecture(city, address, pref_aliases):
    address_cjk, city_cjk = compact(address), compact(city)
    address_latin, city_latin = latin_compact(address), latin_compact(city)
    for text in (address_cjk, address_latin, city_cjk, city_latin):
        hits = [(text.find(alias), -len(alias), pref) for alias, pref in pref_aliases.items() if alias and text.find(alias) >= 0]
        if hits:
            return min(hits)[2]
    return None


def parse_municipality(course_id, city, address, muni_by_key, muni_by_pref, corrections, pref_aliases):
    correction = corrections.get(course_id)
    pref = detect_prefecture(city, address, pref_aliases)
    if not pref:
        raise ValueError("无法识别都道府县")
    if correction and correction[0] == pref and tuple(correction) in muni_by_key:
        return tuple(correction)
    cjk, latin = compact(address), latin_compact(address)
    cjk_hits, latin_hits = [], []
    for item in muni_by_pref[pref]:
        for alias in item["aliases"]:
            normalized = compact(alias)
            if normalized and re.search(r"[^\x00-\x7f]", normalized):
                pos = cjk.find(normalized)
                if pos >= 0:
                    cjk_hits.append((pos, -len(normalized), item["name"]))
        roman = latin_compact(item["roman"])
        if roman:
            pos = latin.find(roman)
            if pos >= 0:
                latin_hits.append((pos, -len(roman), item["name"]))
    hits = cjk_hits or latin_hits
    if not hits:
        raise ValueError(f"已识别 {pref}，但无法识别市町村")
    return pref, min(hits)[2]


def pick(labels, *languages):
    for language in languages:
        if labels.get(language):
            return labels[language]
    return ""


def clean_label(label, language=""):
    label = re.sub(r"\s*[,，].*$", "", label or "")
    label = re.sub(r"\s*[（(].*[）)]$", "", label)
    label = re.sub(r"\s+(City|Town|Village|Ward)$", "", label, flags=re.I)
    label = re.sub(r"[-\s](shi|cho|chō|machi|mura|son|ku)$", "", label, flags=re.I)
    if language == "ko":
        label = re.sub(r"(특별자치)?[시군구정촌]$", "", label)
    return label


def transliterate(kana, table, nchar):
    output, index = [], 0
    while index < len(kana):
        pair = kana[index:index + 2]
        if pair in table:
            output.append(table[pair]); index += 2; continue
        char = kana[index]
        if char == "ン":
            output.append(nchar)
        elif char not in {"ッ", "ー"}:
            output.append(table.get(char, ""))
        index += 1
    return "".join(output)


def municipality_names(item):
    labels, english = item["labels"], clean_label(pick(item["labels"], "en") or item["roman"])
    names = {
        "zh_cn": item["zh_cn"], "zh_tw": item["zh_tw"], "en": english,
        "ko": clean_label(pick(labels, "ko") or transliterate(item["kana"], KO, "ㄴ"), "ko"),
        "th": clean_label(pick(labels, "th") or transliterate(item["kana"], TH, "น")),
    }
    for language in LANGUAGES:
        names[language] = clean_label(pick(labels, language) or english, language)
    return names


def prefecture_names(item):
    labels = item["labels"]
    names = {
        "zh_cn": item["zh_cn"], "zh_tw": item["zh_tw"], "en": item["en"],
        "ko": clean_label(pick(labels, "ko") or item["en"], "ko"),
        "th": clean_label(pick(labels, "th") or item["en"]),
    }
    for language in LANGUAGES:
        names[language] = clean_label(pick(labels, language) or item["en"], language)
    return names


def style_sheet(ws, widths):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")


def locate_sheet(workbook, requested=None):
    if requested:
        if requested not in workbook.sheetnames:
            raise ValueError(f"找不到 Sheet：{requested}。现有 Sheet：{', '.join(workbook.sheetnames)}")
        candidates = [workbook[requested]]
    else:
        candidates = workbook.worksheets
    for ws in candidates:
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        if all(column in headers for column in REQUIRED_COLUMNS):
            return ws, {name: headers.index(name) for name in REQUIRED_COLUMNS}
    raise ValueError("没有找到包含以下字段的 Sheet：" + "、".join(REQUIRED_COLUMNS))


def discover_input():
    candidates = [p for p in TOOL_DIR.glob("*.xlsx") if not p.name.startswith("~$") and "多语言对照汇总" not in p.name]
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        raise ValueError("目录中没有可处理的 .xlsx 文件。请把 Excel 拖到“开始处理Excel.bat”上。")
    names = "、".join(p.name for p in candidates)
    raise ValueError(f"目录中有多个原始 Excel（{names}）。请把需要处理的文件拖到“开始处理Excel.bat”上。")


def unused_output_path(path: Path):
    """默认不覆盖之前的结果，存在同名文件时依次添加 _2、_3。"""
    if not path.exists():
        return path
    sequence = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{sequence}{path.suffix}")
        if not candidate.exists():
            return candidate
        sequence += 1


def write_issue_report(path, issues):
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["原始行号", "球场ID", "球场名称", "城市", "地址", "处理结果/原因"])
        writer.writerows(issues)


def build(input_path: Path, output_path: Path, sheet_name=None):
    prefectures, pref_by_ja, muni_by_key, muni_by_pref, corrections = load_reference()
    pref_order = {item["ja"]: index for index, item in enumerate(prefectures)}
    pref_aliases = build_pref_aliases(prefectures)
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    source, columns = locate_sheet(workbook, sheet_name)
    records, issues = [], []
    for row_number, row in enumerate(source.iter_rows(min_row=2, values_only=True), 2):
        values = {name: row[index] if index < len(row) else None for name, index in columns.items()}
        course_id = id_text(values["球场id中文"])
        name, city, address = str(values["球场名称中文"] or "").strip(), values["城市"], values["地址"]
        if not city or not address:
            issues.append([row_number, course_id, name, city or "", address or "", "已过滤：城市或地址为空"])
            continue
        try:
            pref, municipality = parse_municipality(course_id, city, address, muni_by_key, muni_by_pref, corrections, pref_aliases)
        except ValueError as error:
            issues.append([row_number, course_id, name, city, address, f"未导入：{error}"])
            continue
        records.append({"id": course_id, "name": name, "pref": pref, "muni": municipality, "address": str(address)})

    unique = sorted({(r["pref"], r["muni"]) for r in records}, key=lambda key: (pref_order[key[0]], key[1]))
    muni_names = {key: municipality_names(muni_by_key[key]) for key in unique}
    pref_names = {item["ja"]: prefecture_names(item) for item in prefectures}

    result = Workbook()
    detail = result.active
    detail.title = "球场明细表（已修正版）"
    detail.append(["球场ID", "国家", "球场名称", "县/都道府县（中文）", "都道府県（日文）", "提取市町村名（中文）", "提取市町村名（日文）", "提取市町村名（繁体）", "市町村英文名", "市町村韩文名", "市町村泰文名", "原始地址"])
    for record in records:
        names = muni_names[(record["pref"], record["muni"])]
        detail.append([record["id"], "日本", record["name"], pref_by_ja[record["pref"]]["zh_cn"], record["pref"], names["zh_cn"], record["muni"], names["zh_tw"], names["en"], names["ko"], names["th"], record["address"]])
    style_sheet(detail, [12, 9, 34, 20, 18, 22, 22, 22, 20, 20, 22, 58])
    for cell in detail["A"][1:]:
        cell.number_format = "@"

    muni_counts = Counter((r["pref"], r["muni"]) for r in records)
    summary = result.create_sheet("多语言市町村汇总表")
    summary.append(["序号", "都道府县（中文）", "都道府県（日文）", "中文 (带后缀)", "市町村名（日文）", "繁体 (带后缀)", "英文 (English)", "韩文 (한국어)", "泰文 (ไทย)", "德文 (Deutsch)", "法文 (Français)", "西班牙文 (Español)", "瑞典文 (Svenska)", "荷兰文 (Nederlands)", "越南文 (Tiếng Việt)", "马来文 (Bahasa Melayu)", "印尼文 (Bahasa Indonesia)", "包含球场数量"])
    for sequence, key in enumerate(unique, 1):
        pref, municipality = key
        names = muni_names[key]
        summary.append([sequence, pref_by_ja[pref]["zh_cn"], pref, names["zh_cn"], municipality, names["zh_tw"], names["en"], names["ko"], names["th"], names["de"], names["fr"], names["es"], names["sv"], names["nl"], names["vi"], names["ms"], names["id"], muni_counts[key]])
    style_sheet(summary, [9, 20, 18, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22, 16])

    pref_counts = Counter(r["pref"] for r in records)
    pref_sheet = result.create_sheet("都道府县多语言对照表")
    pref_sheet.append(["序号", "都道府县（中文）", "都道府県（日文）", "繁体中文", "英文 (English)", "韩文 (한국어)", "泰文 (ไทย)", "德文 (Deutsch)", "法文 (Français)", "西班牙文 (Español)", "瑞典文 (Svenska)", "荷兰文 (Nederlands)", "越南文 (Tiếng Việt)", "马来文 (Bahasa Melayu)", "印尼文 (Bahasa Indonesia)", "包含球场总数"])
    for sequence, item in enumerate(prefectures, 1):
        names = pref_names[item["ja"]]
        pref_sheet.append([sequence, item["zh_cn"], item["ja"], item["zh_tw"], names["en"], names["ko"], names["th"], names["de"], names["fr"], names["es"], names["sv"], names["nl"], names["vi"], names["ms"], names["id"], pref_counts[item["ja"]]])
    style_sheet(pref_sheet, [9, 20, 18, 20, 20, 20, 20, 22, 22, 22, 22, 22, 22, 24, 24, 16])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if input_path.resolve() == output_path.resolve():
        raise ValueError("输出文件不能覆盖原始输入文件")
    result.save(output_path)
    issue_path = output_path.with_name(output_path.stem + "_异常记录.csv")
    write_issue_report(issue_path, issues)
    return {"source_rows": source.max_row - 1, "imported": len(records), "issues": len(issues), "municipalities": len(unique), "output": output_path, "issue_report": issue_path}


def parse_args():
    parser = argparse.ArgumentParser(description="日本球场地区多语言标准化工具")
    parser.add_argument("input", nargs="?", help="原始 .xlsx 文件路径；不填写时尝试自动寻找")
    parser.add_argument("-o", "--output", help="输出 .xlsx 文件路径")
    parser.add_argument("--sheet", help="指定原始数据所在 Sheet；通常无需填写")
    return parser.parse_args()


def main():
    args = parse_args()
    try:
        input_path = Path(args.input).expanduser().resolve() if args.input else discover_input().resolve()
        if input_path.suffix.lower() != ".xlsx" or not input_path.exists():
            raise ValueError(f"输入文件不存在或不是 .xlsx：{input_path}")
        output_path = Path(args.output).expanduser().resolve() if args.output else unused_output_path(input_path.with_name(input_path.stem + "_日本地区多语言对照汇总.xlsx"))
        report = build(input_path, output_path, args.sheet)
        print("\n处理完成！")
        print(f"原始数据：{report['source_rows']} 行")
        print(f"成功导入：{report['imported']} 行")
        print(f"过滤/待检查：{report['issues']} 行")
        print(f"市町村数量：{report['municipalities']} 个")
        print(f"输出文件：{report['output']}")
        print(f"异常记录：{report['issue_report']}")
        return 0
    except Exception as error:
        print(f"\n处理失败：{error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
